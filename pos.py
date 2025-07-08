import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os

# === 初期設定 ===
folder = "C:/Users/maeka/Documents/pos"
os.makedirs(folder, exist_ok=True)

date_str = datetime.now().strftime("%Y-%m-%d")
log_file = f"{folder}/sales_log_{date_str}.xlsx"
summary_file = f"{folder}/sales_summary_{date_str}.xlsx"
product_master_file = f"{folder}/product_master.xlsx"

# === 商品マスターの読み込み（存在しなければ初期商品） ===
try:
    df_master = pd.read_excel(product_master_file)
    items = dict(zip(df_master["商品名"], df_master["価格"]))
except FileNotFoundError:
    items = {"マドレーヌ": 300, "焼きそば": 250}
    df_master = pd.DataFrame(list(items.items()), columns=["商品名", "価格"])
    df_master.to_excel(product_master_file, index=False)

sales = {i: 0 for i in items}

# === ページ構成 ===
tab1, tab2 = st.tabs(["🛍️ 販売ページ", "🧑‍💼 管理ページ"])

# === 販売ページ ===
with tab1:
    st.markdown("<h1 style='text-align:center;'>🎪 商品販売</h1>", unsafe_allow_html=True)
    item = st.selectbox("販売する商品を選んでください", list(items.keys()))
    st.markdown("### 🔢 販売個数を入力してください")
    count = st.number_input("販売個数", min_value=1, value=1, step=1)

    if st.button("販売する"):
        price = items[item]
        total = price * count
        sales[item] += count
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # ログ保存
        log_df = pd.DataFrame([[timestamp, item, count, price, total]],
                              columns=["販売時刻", "商品名", "販売個数", "単価", "合計金額"])

        try:
            book = load_workbook(log_file)
            with pd.ExcelWriter(log_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                startrow = book["Sheet1"].max_row
                log_df.to_excel(writer, startrow=startrow, index=False, header=False)
        except FileNotFoundError:
            log_df.to_excel(log_file, index=False)

        # 集計保存
        summary_df = pd.DataFrame([[i, sales[i], items[i], sales[i]*items[i]] for i in items],
                                  columns=["商品名", "販売個数", "単価", "合計金額"])
        summary_df.to_excel(summary_file, index=False)

        st.success(f"{item} を {count} 個販売しました！（合計 ¥{total}）")

# === 管理者ページ ===
with tab2:
    st.markdown("<h1 style='text-align:center;'>🧑‍💼 管理者ページ</h1>", unsafe_allow_html=True)
    admin_code = st.text_input("パスコードを入力してください", type="password")

    if admin_code != "kazuki12@":
        st.warning("正しいパスコードを入力してください。")
    else:
        # 商品追加
        with st.expander("🆕 商品追加"):
            new_item = st.text_input("商品名", key="add_name")
            new_price = st.number_input("価格", min_value=1, key="add_price")
            if st.button("追加", key="add_button"):
                if new_item in items:
                    st.error("既に存在する商品です。")
                else:
                    items[new_item] = new_price
                    sales[new_item] = 0
                    st.success(f"{new_item} を ￥{new_price} で追加しました。")

                    # Excelにも追記
                    new_df = pd.DataFrame([[new_item, new_price]], columns=["商品名", "価格"])
                    try:
                        book = load_workbook(product_master_file)
                        with pd.ExcelWriter(product_master_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                            startrow = book["Sheet1"].max_row
                            new_df.to_excel(writer, startrow=startrow, index=False, header=False)
                    except FileNotFoundError:
                        new_df.to_excel(product_master_file, index=False)

        # 価格変更
        with st.expander("✏️ 価格変更"):
            change_item = st.selectbox("対象商品", list(items.keys()), key="change_item")
            new_price2 = st.number_input("新しい価格", min_value=1, key="change_price")
            if st.button("変更", key="apply_price"):
                items[change_item] = new_price2
                st.success(f"{change_item} の価格を ￥{new_price2} に変更しました。")

                # Excelも更新
                try:
                    df = pd.read_excel(product_master_file)
                    df.loc[df["商品名"] == change_item, "価格"] = new_price2
                    df.to_excel(product_master_file, index=False)
                except FileNotFoundError:
                    pass

        # 商品削除
        with st.expander("🗑️ 商品削除"):
            delete_item = st.selectbox("削除対象", list(items.keys()), key="delete_item")
            if st.button("削除", key="apply_delete"):
                if sales[delete_item] == 0:
                    items.pop(delete_item)
                    sales.pop(delete_item)
                    st.warning(f"{delete_item} を削除しました。")

                    # Excelからも削除
                    try:
                        df = pd.read_excel(product_master_file)
                        df = df[df["商品名"] != delete_item]
                        df.to_excel(product_master_file, index=False)
                    except FileNotFoundError:
                        pass
                else:
                    st.error("販売済みの商品は削除できません。")

        # 販売リセット
        with st.expander("🔄 販売記録リセット"):
            if st.button("リセット", key="reset_sales"):
                for i in sales:
                    sales[i] = 0
                empty_df = pd.DataFrame(columns=["商品名", "販売個数", "単価", "合計金額"])
                empty_df.to_excel(summary_file, index=False)
                st.warning("販売記録を初期化しました。")